import openpyxl
import sys
import os

def readExel():
    filedir = os.path.dirname(os.path.realpath(__file__))
    filename = filedir + r'\1.xlsx'
    wfname = filedir + r'\1.html'
    inwb = openpyxl.load_workbook(filename)  # 读文件

    sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容

    # 获取sheet的最大行数和列数
    rows = ws.max_row+1
    cols = ws.max_column+1
    table = '<table class="zp-recruit_c_date" cellpadding="0" cellspacing="0">'
    for r in range(1,rows):
        table += "<tr>"
        for c in range(1,cols):
            if r==1:
                table += "<th>"+ws.cell(r,c).value+"</th>"
            else:
                table += "<td>"+ws.cell(r,c).value+"</td>"
        table += "</tr>"
    table += '</table>'

    file_handle = open(wfname, mode='w', encoding='utf-8')
    file_handle.write(table)
    file_handle.close()


readExel()
